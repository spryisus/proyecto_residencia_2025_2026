import io
import os
import logging
from datetime import datetime
from typing import List, Dict, Any

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import Response, JSONResponse
import openpyxl

app = FastAPI(title="Excel Generator Service")

ROOT = os.path.dirname(__file__)
TEMPLATES_DIR = os.path.join(ROOT, "assets", "templates")
TEMPLATE_PATH_JUMPERS = os.path.join(TEMPLATES_DIR, "plantilla_jumpers.xlsx")

LAST_GENERATED_FILE_CONTENT: bytes | None = None
LAST_GENERATED_FILENAME: str | None = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _ensure_template(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Template not found: {path}")


def _save_workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@app.get("/", tags=["root"])
def root():
    return {"ok": True, "endpoints": ["/api/generate-jumpers-excel", "/api/debug-last-file"]}


@app.get("/health", tags=["health"])
def health():
    try:
        exists = os.path.exists(TEMPLATE_PATH_JUMPERS)
        return {"ok": True, "templates": {"jumpers": exists}}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.post("/api/generate-jumpers-excel")
async def generate_jumpers_excel(request: Request):
    payload = await request.json()
    items: List[Dict[str, Any]] = payload.get("items") or []
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items must be a non-empty list")

    try:
        _ensure_template(TEMPLATE_PATH_JUMPERS)
        wb = openpyxl.load_workbook(TEMPLATE_PATH_JUMPERS)
        ws = wb.active

        start_row = 2
        for idx, item in enumerate(items, start=0):
            row = start_row + idx
            # Example mapping: adjust according to actual template
            ws.cell(row=row, column=1, value=item.get("id", ""))
            ws.cell(row=row, column=2, value=item.get("name", ""))
            ws.cell(row=row, column=3, value=item.get("location", ""))
            ws.cell(row=row, column=4, value=item.get("status", ""))

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_jumpers_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except FileNotFoundError as e:
        logger.exception("Template missing")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("Error generating jumpers excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/debug-last-file")
def debug_last_file():
    if not LAST_GENERATED_FILE_CONTENT or not LAST_GENERATED_FILENAME:
        raise HTTPException(status_code=404, detail="No generated file in memory")

    tmp_path = os.path.join("/tmp", LAST_GENERATED_FILENAME)
    try:
        with open(tmp_path, "wb") as f:
            f.write(LAST_GENERATED_FILE_CONTENT)
        size = os.path.getsize(tmp_path)
        return {"ok": True, "path": tmp_path, "size": size}
    except Exception as e:
        logger.exception("Failed to write debug file")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("excel_generator_service.main:app", host="0.0.0.0", port=8001, reload=True)
